#!/usr/bin/env python3
"""
BSC链内盘高胜率钱包分析器
数据源: GMGN.ai
功能: 爬取BSC链上Meme币交易的高胜率钱包，生成XLSX报表
"""

import requests
import time
import json
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
import sys
import os

# ============================================================
# 配置区域 - 可根据需要调整
# ============================================================
CONFIG = {
    "chain": "bsc",                    # 链: bsc
    "min_win_rate": 50,                # 最低胜率筛选 (%)
    "min_trades": 5,                   # 最少交易次数
    "top_wallets": 100,                # 获取前N个钱包
    "timeframes": ["1d", "7d", "30d"], # 分析的时间段
    "output_file": "BSC_高胜率钱包分析.xlsx",
    "request_delay": 1.5,              # 请求间隔(秒)，避免被限流
}

# GMGN API 基础配置
GMGN_BASE = "https://gmgn.ai"
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                  "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Accept": "application/json, text/plain, */*",
    "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8",
    "Referer": "https://gmgn.ai/",
    "Origin": "https://gmgn.ai",
}

SESSION = requests.Session()
SESSION.headers.update(HEADERS)


# ============================================================
# 数据获取模块
# ============================================================

def fetch_top_traders(timeframe="7d", orderby="winrate", limit=50):
    """
    获取GMGN上BSC链的顶级交易者(钱包)排名
    endpoint: /defi/quotation/v1/rank/{chain}/wallets/{timeframe}
    orderby: winrate / realized_profit_7d / pnl_7d / profit
    """
    url = f"{GMGN_BASE}/defi/quotation/v1/rank/{CONFIG['chain']}/wallets/{timeframe}"
    params = {
        "orderby": orderby,
        "direction": "desc",
        "limit": limit,
    }

    print(f"  [请求] 获取 {timeframe} 钱包排行榜 (排序: {orderby}, 数量: {limit})...")
    try:
        resp = SESSION.get(url, params=params, timeout=30)
        resp.raise_for_status()
        data = resp.json()

        if data.get("code") == 0 and data.get("data"):
            rank_data = data["data"].get("rank", [])
            print(f"  [成功] 获取到 {len(rank_data)} 个钱包")
            return rank_data
        else:
            print(f"  [警告] API返回异常: {data.get('msg', '未知错误')}")
            return []
    except requests.exceptions.RequestException as e:
        print(f"  [错误] 请求失败: {e}")
        return []


def fetch_wallet_detail(wallet_address, period="7d"):
    """获取单个钱包的详细交易统计"""
    url = f"{GMGN_BASE}/defi/quotation/v1/smartmoney/{CONFIG['chain']}/walletNew/{wallet_address}"
    params = {"period": period}

    try:
        resp = SESSION.get(url, params=params, timeout=30)
        resp.raise_for_status()
        data = resp.json()
        if data.get("code") == 0 and data.get("data"):
            return data["data"]
        return None
    except requests.exceptions.RequestException:
        return None


def fetch_token_info(token_address):
    """获取代币基本信息"""
    url = f"{GMGN_BASE}/api/v1/token_info/{CONFIG['chain']}/{token_address}"
    try:
        resp = SESSION.get(url, timeout=15)
        resp.raise_for_status()
        data = resp.json()
        if data.get("code") == 0 and data.get("data"):
            return data["data"]
        return None
    except requests.exceptions.RequestException:
        return None


def fetch_token_top_buyers(token_address):
    """获取代币的前70个早期买家"""
    url = f"{GMGN_BASE}/defi/quotation/v1/tokens/top_buyers/{CONFIG['chain']}/{token_address}"
    try:
        resp = SESSION.get(url, timeout=30)
        resp.raise_for_status()
        data = resp.json()
        if data.get("code") == 0 and data.get("data"):
            holders_data = data["data"].get("holders", {})
            holder_info = holders_data.get("holderInfo", [])
            return holder_info, holders_data
        return [], {}
    except requests.exceptions.RequestException:
        return [], {}


def fetch_token_trades(token_address, event="buy", limit=200):
    """
    获取代币的交易记录，提取买入钱包
    通过分页获取更多交易记录
    """
    all_trades = []
    cursor = None
    fetched = 0

    while fetched < limit:
        batch_size = min(50, limit - fetched)
        url = f"{GMGN_BASE}/api/v1/token_trades/{CONFIG['chain']}/{token_address}"
        params = {"limit": batch_size, "event": event}
        if cursor:
            params["cursor"] = cursor

        try:
            resp = SESSION.get(url, params=params, timeout=30)
            resp.raise_for_status()
            data = resp.json()
            history = data.get("data", {}).get("history", [])
            if not history:
                break
            all_trades.extend(history)
            fetched += len(history)

            # 获取下一页 cursor
            next_cursor = data.get("data", {}).get("next")
            if next_cursor and next_cursor != cursor:
                cursor = next_cursor
            else:
                break

            time.sleep(CONFIG["request_delay"])
        except requests.exceptions.RequestException:
            break

    return all_trades


def fetch_wallet_stats(wallet_address):
    """
    获取单个钱包的交易统计 (walletNew endpoint)
    返回: realized_profit_7d, pnl_7d, buy_7d, sell_7d, tags, twitter 等
    注意: winrate 字段在此endpoint始终为None，需从排行榜获取
    """
    url = f"{GMGN_BASE}/defi/quotation/v1/smartmoney/{CONFIG['chain']}/walletNew/{wallet_address}"
    params = {"period": "7d"}
    try:
        resp = SESSION.get(url, params=params, timeout=15)
        resp.raise_for_status()
        data = resp.json()
        if data.get("code") == 0 and data.get("data"):
            return data["data"]
        return None
    except requests.exceptions.RequestException:
        return None


def estimate_winrate(detail):
    """
    根据walletNew数据估算胜率。
    GMGN免费API的walletNew不返回winrate字段，
    用 all_pnl, total_profit_pnl, realized_profit 等指标进行估算。
    返回: (estimated_winrate_pct, confidence)
    """
    if not detail:
        return 0, "无"

    all_pnl = safe_float(detail.get("all_pnl", 0))
    realized_30d = safe_float(detail.get("realized_profit_30d", 0))
    pnl_30d = safe_float(detail.get("pnl_30d", 0))
    buy_30d = safe_int(detail.get("buy_30d", 0))
    sell_30d = safe_int(detail.get("sell_30d", 0))
    total_value = safe_float(detail.get("total_value", 0))

    if buy_30d + sell_30d < 5:
        return 0, "无"

    scores = []

    # 维度1: 总体PNL (权重最高)
    if all_pnl > 0.05: scores.append(75)
    elif all_pnl > 0.01: scores.append(60)
    elif all_pnl > 0: scores.append(52)
    elif all_pnl > -0.01: scores.append(45)
    elif all_pnl > -0.05: scores.append(35)
    else: scores.append(20)

    # 维度2: 30天PNL
    if pnl_30d > 0.05: scores.append(72)
    elif pnl_30d > 0.01: scores.append(58)
    elif pnl_30d > 0: scores.append(52)
    elif pnl_30d > -0.02: scores.append(42)
    else: scores.append(25)

    # 维度3: 已实现利润
    if realized_30d > 10000: scores.append(78)
    elif realized_30d > 1000: scores.append(65)
    elif realized_30d > 0: scores.append(55)
    elif realized_30d > -1000: scores.append(40)
    else: scores.append(25)

    # 维度4: 资产规模
    if total_value > 100000: scores.append(68)
    elif total_value > 10000: scores.append(58)
    elif total_value > 1000: scores.append(50)
    else: scores.append(38)

    weights = [3, 2.5, 2, 1]
    estimated = sum(s * w for s, w in zip(scores, weights)) / sum(weights)

    if buy_30d + sell_30d >= 50 and total_value > 5000:
        confidence = "高"
    elif buy_30d + sell_30d >= 20:
        confidence = "中"
    else:
        confidence = "低"

    return round(estimated, 1), confidence


def collect_wallet_data_by_ca(token_address):
    """
    根据代币合约地址(CA)收集该代币的买入者钱包，并分析其内盘交易表现。

    策略(v5 - 内盘购买数据 + 胜率筛选):
      1. 获取代币信息
      2. 获取所有买入记录 (token_trades) - 含每笔买入金额/价格/该币盈亏
      3. 获取前70早期买家补充 (top_buyers)
      4. 加载排行榜钱包 (精确胜率)
      5. 按钱包聚合内盘购买数据，附加胜率
      6. 按胜率筛选输出
    """
    print("\n" + "=" * 60)
    print(f"  按合约地址(CA)分析内盘买入者 (v5 内盘购买数据)")
    print("=" * 60)

    # Step 1: 获取代币信息
    print(f"\n▶ Step 1/6 获取代币信息: {token_address[:10]}...")
    token_info = fetch_token_info(token_address)
    if token_info:
        symbol = token_info.get("symbol", "UNKNOWN")
        name = token_info.get("name", "")
        print(f"  代币: {name} ({symbol})")
        print(f"  持有人数: {token_info.get('holder_count', 'N/A')}")
    else:
        symbol = "UNKNOWN"
        name = ""
        print("  [警告] 无法获取代币信息，继续尝试...")
    time.sleep(0.8)

    # Step 2: 获取所有买入交易记录 (核心数据源)
    print(f"\n▶ Step 2/6 获取内盘买入交易记录...")
    trades = fetch_token_trades(token_address, event="buy", limit=500)
    print(f"  获取到 {len(trades)} 条买入记录")

    # 按钱包聚合: 每个钱包的买入数据
    wallet_trades = {}
    for trade in trades:
        addr = trade.get("maker", "")
        if not addr:
            continue
        if addr not in wallet_trades:
            wallet_trades[addr] = {
                "trades": [],
                "tags": trade.get("maker_tags", []) or [],
                "maker_token_tags": trade.get("maker_token_tags", []) or [],
                "twitter_name": trade.get("maker_twitter_name", ""),
                "twitter_username": trade.get("maker_twitter_username", ""),
                "name": trade.get("maker_name", ""),
            }
        wallet_trades[addr]["trades"].append(trade)

    print(f"  不重复买入钱包: {len(wallet_trades)} 个")

    # Step 3: 获取前70早期买家 (补充状态信息)
    print(f"\n▶ Step 3/6 获取前70早期买家...")
    top_buyers, holders_summary = fetch_token_top_buyers(token_address)
    top_buyer_map = {}
    for buyer in top_buyers:
        addr = buyer.get("wallet_address", "")
        if addr:
            top_buyer_map[addr] = {
                "status": buyer.get("status", ""),
                "tags": buyer.get("tags", []) or [],
                "maker_token_tags": buyer.get("maker_token_tags", []) or [],
                "is_top70": True,
            }
            # 如果top_buyers里有但token_trades里没有的钱包，也加入
            if addr not in wallet_trades:
                wallet_trades[addr] = {
                    "trades": [],
                    "tags": buyer.get("tags", []) or [],
                    "maker_token_tags": buyer.get("maker_token_tags", []) or [],
                    "twitter_name": "",
                    "twitter_username": "",
                    "name": "",
                }
    print(f"  早期买家: {len(top_buyer_map)} 个")
    print(f"  总计钱包: {len(wallet_trades)} 个")
    time.sleep(0.8)

    # Step 4: 加载排行榜钱包 (精确胜率)
    print(f"\n▶ Step 4/6 加载排行榜数据 (精确胜率来源)...")
    ranked_wallets = {}
    orderbys = ["winrate", "pnl", "txs", "volume", "buy", "sell", "net_inflow"]
    timeframes = ["1d", "7d", "30d"]

    for tf in timeframes:
        for ob in orderbys:
            for direction in ["desc", "asc"]:
                traders = fetch_top_traders(timeframe=tf, orderby=ob, limit=200)
                for t in traders:
                    addr = t.get("wallet_address", "") or t.get("address", "")
                    if addr and addr not in ranked_wallets:
                        wr = safe_float(t.get("winrate_7d") or t.get("winrate_30d") or t.get("winrate_1d", 0))
                        if wr > 0:
                            ranked_wallets[addr] = t
                time.sleep(0.3)
    print(f"  排行榜钱包池 (有精确胜率): {len(ranked_wallets)} 个")

    # Step 5: 只查前70早期买家的胜率数据 (跳过其他钱包节省时间)
    top70_addrs = set(top_buyer_map.keys())
    total_to_query = len(top70_addrs)
    print(f"\n▶ Step 5/6 查询 {total_to_query} 个前70早期买家的胜率数据...")

    wallet_results = []
    precise_count = 0
    estimated_count = 0
    no_data_count = 0

    query_idx = 0
    for addr, wt in wallet_trades.items():
        top_info = top_buyer_map.get(addr, {})
        is_top70 = top_info.get("is_top70", False)

        # 只处理前70早期买家
        if not is_top70:
            continue

        query_idx += 1
        if query_idx % 10 == 0 or query_idx == 1:
            print(f"  进度: {query_idx}/{total_to_query}")

        # --- 聚合该钱包在此代币上的内盘数据 ---
        trade_list = wt["trades"]
        top_info = top_buyer_map.get(addr, {})

        if trade_list:
            # 有交易记录: 从最新一条获取该钱包在此币的汇总数据
            latest = trade_list[0]  # 最新交易
            earliest = trade_list[-1]  # 最早交易

            first_buy_time = min(safe_int(t.get("timestamp", 0)) for t in trade_list)
            total_buy_usd = sum(safe_float(t.get("amount_usd", 0)) for t in trade_list)
            total_buy_count = len(trade_list)

            # 最新交易包含该钱包在此币的累计数据
            token_balance = safe_float(latest.get("balance", 0))
            history_bought = safe_float(latest.get("history_bought_amount", 0))
            history_sold_income = safe_float(latest.get("history_sold_income", 0))
            history_sold_amount = safe_float(latest.get("history_sold_amount", 0))
            realized_profit = safe_float(latest.get("realized_profit", 0))
            unrealized_profit = safe_float(latest.get("unrealized_profit", 0))
            total_trade_on_token = safe_int(latest.get("total_trade", 0))
            buy_price = safe_float(earliest.get("price_usd", 0))  # 首次买入价
            latest_price = safe_float(latest.get("price_usd", 0))  # 最新买入价
        else:
            # 仅在top_buyers中，无交易记录
            first_buy_time = 0
            total_buy_usd = 0
            total_buy_count = 0
            token_balance = 0
            history_bought = 0
            history_sold_income = 0
            history_sold_amount = 0
            realized_profit = 0
            unrealized_profit = 0
            total_trade_on_token = 0
            buy_price = 0
            latest_price = 0

        # 格式化首次买入时间
        if first_buy_time:
            try:
                first_buy_str = datetime.fromtimestamp(int(first_buy_time)).strftime("%Y-%m-%d %H:%M")
            except:
                first_buy_str = "N/A"
        else:
            first_buy_str = "N/A"

        # 买入状态
        buyer_status = top_info.get("status", "")
        if not buyer_status:
            if token_balance > 0 and history_sold_amount == 0:
                buyer_status = "hold"
            elif token_balance > 0 and history_sold_amount > 0:
                buyer_status = "sold_part"
            elif token_balance == 0 and history_sold_amount > 0:
                buyer_status = "sold"
            else:
                buyer_status = "unknown"

        is_top70 = top_info.get("is_top70", False)

        # --- 获取胜率 ---
        has_ranking = addr in ranked_wallets
        win_rate = 0
        winrate_source = "无数据"
        winrate_confidence = ""

        if has_ranking:
            ranked_data = ranked_wallets[addr]
            wr_raw = safe_float(ranked_data.get("winrate_7d") or ranked_data.get("winrate_30d") or ranked_data.get("winrate_1d", 0))
            win_rate = wr_raw * 100 if wr_raw <= 1 else wr_raw
            winrate_source = "排行榜"
            winrate_confidence = "精确"
            precise_count += 1
        else:
            # 查walletNew估算胜率
            detail = fetch_wallet_stats(addr)
            time.sleep(0.4)
            if detail:
                est_wr, confidence = estimate_winrate(detail)
                win_rate = est_wr
                if est_wr > 0:
                    winrate_source = "估算"
                    winrate_confidence = confidence
                    estimated_count += 1
                else:
                    no_data_count += 1
            else:
                no_data_count += 1

        # 合并标签
        all_tags = list(set(
            (wt.get("tags", []) or []) +
            (top_info.get("tags", []) or []) +
            (wt.get("maker_token_tags", []) or []) +
            (top_info.get("maker_token_tags", []) or [])
        ))

        twitter = wt.get("twitter_name") or wt.get("twitter_username") or ""

        wallet_results.append({
            "address": addr,
            "short_address": f"{addr[:6]}...{addr[-4:]}",
            "win_rate": win_rate,
            "winrate_source": winrate_source,
            "winrate_confidence": winrate_confidence,
            # 内盘购买数据 (该代币)
            "first_buy_time": first_buy_str,
            "first_buy_timestamp": first_buy_time,
            "total_buy_usd": total_buy_usd,
            "buy_price": buy_price,
            "latest_price": latest_price,
            "total_buy_count": total_buy_count,
            "total_trade_on_token": total_trade_on_token,
            "token_balance": token_balance,
            "history_bought": history_bought,
            "history_sold_income": history_sold_income,
            "realized_profit": realized_profit,
            "unrealized_profit": unrealized_profit,
            "total_profit_token": realized_profit + unrealized_profit,
            # 状态和标签
            "buyer_status": buyer_status,
            "is_top70": is_top70,
            "tags": ", ".join(all_tags) if all_tags else "",
            "twitter": twitter,
            "gmgn_link": f"https://gmgn.ai/bsc/address/{addr}",
        })

    # Step 6: 只保留前70早期买家 + 按胜率筛选排序
    min_wr = CONFIG["min_win_rate"]
    print(f"\n▶ Step 6/6 筛选前70早期买家中胜率 ≥ {min_wr}% 的钱包...")

    # 只保留前70早期买家
    top70_results = [w for w in wallet_results if w.get("is_top70")]
    print(f"  前70早期买家: {len(top70_results)} 个")

    # 达标钱包: 精确胜率优先，然后估算
    precise_ok = [w for w in top70_results if w["winrate_source"] == "排行榜" and w["win_rate"] >= min_wr]
    precise_ok.sort(key=lambda x: (-x["win_rate"], -x["total_buy_usd"]))

    estimated_ok = [w for w in top70_results if w["winrate_source"] == "估算" and w["win_rate"] >= min_wr]
    estimated_ok.sort(key=lambda x: (-x["win_rate"], -x["total_buy_usd"]))

    # 低于阈值但有胜率的 (参考)
    below = [w for w in top70_results if w["win_rate"] > 0 and w["win_rate"] < min_wr]
    below.sort(key=lambda x: (-x["win_rate"], -x["total_buy_usd"]))

    qualified = precise_ok + estimated_ok + below

    print(f"\n  ========== 分析结果 ==========")
    print(f"  总买入者: {len(wallet_trades)} 个")
    print(f"  前70早期买家: {len(top70_results)} 个")
    print(f"  排行榜精确胜率: {precise_count} 个 (达标≥{min_wr}%: {len(precise_ok)})")
    print(f"  估算胜率: {estimated_count} 个 (达标≥{min_wr}%: {len(estimated_ok)})")
    print(f"  无法估算: {no_data_count} 个")
    print(f"  ---")
    print(f"  高胜率钱包 (≥{min_wr}%): {len(precise_ok) + len(estimated_ok)} 个")
    print(f"  输出总计: {len(qualified)} 个钱包")

    return qualified, token_info


def parse_ca_wallet_data(address, detail, meta, token_address, token_symbol):
    """解析CA模式下的钱包数据"""
    try:
        # 胜率
        winrate_raw = detail.get("winrate") or detail.get("winrate_7d") or 0
        win_rate = safe_float(winrate_raw)
        if win_rate and win_rate <= 1:
            win_rate = win_rate * 100

        # 交易统计
        buy_7d = safe_int(detail.get("buy_7d") or detail.get("buy", 0))
        sell_7d = safe_int(detail.get("sell_7d") or detail.get("sell", 0))
        total_trades = buy_7d + sell_7d

        # 利润
        realized = safe_float(detail.get("realized_profit_7d") or detail.get("realized_profit", 0))
        unrealized = safe_float(detail.get("unrealized_profit", 0))
        total_profit = realized + unrealized

        pnl_7d = safe_float(detail.get("pnl_7d", 0))
        pnl_30d = safe_float(detail.get("pnl_30d", 0))

        # 余额
        balance = safe_float(detail.get("bnb_balance") or detail.get("eth_balance") or detail.get("balance", 0))
        total_value = safe_float(detail.get("total_value", 0))

        # 标签
        tags_from_detail = detail.get("tags", [])
        tags_from_meta = meta.get("tags", [])
        all_tags = list(set(
            (tags_from_detail if isinstance(tags_from_detail, list) else []) +
            (tags_from_meta if isinstance(tags_from_meta, list) else [])
        ))

        # 社交信息
        twitter = detail.get("twitter_name") or detail.get("twitter_username") or \
                  meta.get("twitter_name", "") or meta.get("twitter_username", "") or ""

        # 持仓时间
        avg_hold = detail.get("avg_holding_peroid") or detail.get("avg_holding_period_7d", "N/A")

        # 最后活跃
        last_active = detail.get("last_active_timestamp", 0)
        if last_active:
            last_active_str = datetime.fromtimestamp(int(last_active)).strftime("%Y-%m-%d %H:%M")
        else:
            last_active_str = "N/A"

        # PNL 分布
        win_count = (
            safe_int(detail.get("pnl_lt_2x_num", 0))
            + safe_int(detail.get("pnl_2x_5x_num", 0))
            + safe_int(detail.get("pnl_gt_5x_num", 0))
        )
        loss_count = (
            safe_int(detail.get("pnl_lt_minus_dot5_num", 0))
            + safe_int(detail.get("pnl_minus_dot5_0x_num", 0))
        )

        # 来源信息
        source = meta.get("source", "")
        buyer_status = meta.get("status", "")
        maker_token_tags = meta.get("maker_token_tags", [])

        # 该代币上的利润
        token_realized = safe_float(meta.get("realized_profit", 0))
        token_unrealized = safe_float(meta.get("unrealized_profit", 0))

        # 是否合约
        is_contract = detail.get("is_contract", False)

        return {
            "address": address,
            "short_address": f"{address[:6]}...{address[-4:]}",
            "win_rate": win_rate,
            "total_trades": total_trades,
            "buy_count": buy_7d,
            "sell_count": sell_7d,
            "win_count": win_count,
            "loss_count": loss_count,
            "realized_profit_usd": realized,
            "unrealized_profit_usd": unrealized,
            "total_profit_usd": total_profit,
            "pnl_7d": pnl_7d,
            "pnl_30d": pnl_30d,
            "roi_percent": pnl_7d * 100 if pnl_7d else 0,
            "avg_hold_time": format_hold_time(avg_hold),
            "max_single_profit": 0,
            "token_count": 0,
            "last_active": last_active_str,
            "tags": ", ".join(all_tags) if all_tags else "",
            "twitter": twitter,
            "follow_count": safe_int(detail.get("followers_count", 0)),
            "volume_7d": 0,
            "balance_bnb": balance,
            "total_value_usd": total_value,
            "pnl_gt_5x_count": safe_int(detail.get("pnl_gt_5x_num", 0)),
            "pnl_2x_5x_count": safe_int(detail.get("pnl_2x_5x_num", 0)),
            "buyer_status": buyer_status,
            "source": source,
            "maker_token_tags": ", ".join(maker_token_tags) if isinstance(maker_token_tags, list) else "",
            "token_realized_profit": token_realized,
            "token_unrealized_profit": token_unrealized,
            "is_contract": is_contract,
            "gmgn_link": f"https://gmgn.ai/bsc/address/{address}",
        }
    except Exception as e:
        print(f"\n  [解析错误] {address[:10]}...: {e}")
        return None


def generate_ca_xlsx(wallets, token_info, token_address, filename):
    """生成CA分析模式的 XLSX 报表"""
    if not wallets:
        print("\n⚠ 没有获取到钱包数据")
        return None

    wb = Workbook()

    symbol = token_info.get("symbol", "UNKNOWN") if token_info else "UNKNOWN"
    name = token_info.get("name", "") if token_info else ""

    # --- Sheet 1: 买入者胜率排行 ---
    create_ca_overview_sheet(wb, wallets, symbol, name, token_address, token_info)

    # --- Sheet 2: 统计 ---
    create_stats_sheet(wb, [w for w in wallets if w["win_rate"] > 0])

    # --- Sheet 3: 使用说明 ---
    create_help_sheet(wb)

    filepath = os.path.join(os.path.dirname(os.path.abspath(__file__)), filename)
    wb.save(filepath)
    print(f"\n✔ 报表已保存: {filepath}")
    return filepath


def create_ca_overview_sheet(wb, wallets, symbol, name, token_address, token_info):
    """创建CA模式的主数据表"""
    ws = wb.active
    ws.title = f"{symbol} 买入者胜率分析"

    # 样式
    title_font = Font(name="微软雅黑", size=16, bold=True, color="FFFFFF")
    title_fill = PatternFill(start_color="1A237E", end_color="1A237E", fill_type="solid")
    header_font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="283593", end_color="283593", fill_type="solid")
    data_font = Font(name="Consolas", size=10)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="E0E0E0"),
        right=Side(style="thin", color="E0E0E0"),
        top=Side(style="thin", color="E0E0E0"),
        bottom=Side(style="thin", color="E0E0E0"),
    )

    def get_winrate_style(rate):
        if rate >= 80:
            return Font(name="Consolas", size=11, bold=True, color="1B5E20"), \
                   PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
        elif rate >= 65:
            return Font(name="Consolas", size=11, bold=True, color="33691E"), \
                   PatternFill(start_color="DCEDC8", end_color="DCEDC8", fill_type="solid")
        elif rate >= 50:
            return Font(name="Consolas", size=10, color="F57F17"), \
                   PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
        elif rate > 0:
            return Font(name="Consolas", size=10, color="B71C1C"), \
                   PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
        else:
            return data_font, PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

    def get_profit_font(value):
        if value > 0:
            return Font(name="Consolas", size=10, color="1B5E20")
        elif value < 0:
            return Font(name="Consolas", size=10, color="B71C1C")
        return data_font

    # 表头定义 - 精简版
    headers = [
        ("排名", 6),
        ("完整地址", 44),
        ("胜率 %", 10),
        ("胜率等级", 10),
        ("置信度", 8),
        ("标签", 22),
        ("GMGN链接", 44),
    ]

    col_count = len(headers)
    last_col = get_column_letter(col_count)

    # 标题行
    ws.merge_cells(f"A1:{last_col}1")
    title_cell = ws["A1"]
    title_cell.value = (
        f"{name} ({symbol}) 前70早期买家胜率筛选 | "
        f"CA: {token_address[:10]}...{token_address[-6:]} | "
        f"{datetime.now().strftime('%Y-%m-%d %H:%M')}"
    )
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    # 摘要行
    ws.merge_cells(f"A2:{last_col}2")
    summary = ws["A2"]
    wallets_with_wr = [w for w in wallets if w["win_rate"] > 0]
    precise_wr = [w for w in wallets if w.get("winrate_source") == "排行榜"]
    estimated_wr = [w for w in wallets if w.get("winrate_source") == "估算"]
    high_wr = [w for w in wallets if w["win_rate"] >= CONFIG["min_win_rate"]]
    avg_wr = sum(w["win_rate"] for w in wallets_with_wr) / len(wallets_with_wr) if wallets_with_wr else 0
    holder_count = token_info.get("holder_count", "N/A") if token_info else "N/A"
    summary.value = (
        f"筛选结果: {len(wallets)} 个钱包 | "
        f"精确胜率(排行榜): {len(precise_wr)} | "
        f"估算胜率: {len(estimated_wr)} | "
        f"高胜率(≥{CONFIG['min_win_rate']}%): {len(high_wr)} | "
        f"平均胜率: {avg_wr:.1f}%"
    )
    summary.font = Font(name="微软雅黑", size=10, color="616161")
    summary.fill = PatternFill(start_color="E8EAF6", end_color="E8EAF6", fill_type="solid")
    summary.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 25

    # 表头行
    for col_idx, (header_text, width) in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=header_text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[3].height = 30

    # 数据行
    for row_idx, w in enumerate(wallets, 4):
        rank = row_idx - 3

        row_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        if rank % 2 == 0:
            row_fill = PatternFill(start_color="FAFAFA", end_color="FAFAFA", fill_type="solid")

        wr_display = w["win_rate"] if w["win_rate"] > 0 else "N/A"
        wr_level = get_winrate_level(w["win_rate"]) if w["win_rate"] > 0 else "无数据"

        row_data = [
            rank,                                      # A: 排名
            w["address"],                              # B: 完整地址
            wr_display,                                # C: 胜率 %
            wr_level,                                  # D: 胜率等级
            w.get("winrate_confidence", ""),            # E: 置信度
            w.get("tags", ""),                         # F: 标签
            w.get("gmgn_link", ""),                    # G: GMGN链接
        ]

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = center_align
            cell.border = thin_border
            cell.fill = row_fill
            cell.font = data_font

        # 胜率列 (C列)
        wr_cell = ws.cell(row=row_idx, column=3)
        if w["win_rate"] > 0:
            wr_font, wr_fill = get_winrate_style(w["win_rate"])
            wr_cell.font = wr_font
            wr_cell.fill = wr_fill
            wr_cell.number_format = '0.0"%"'

            # 胜率等级列 (D列)
            level_cell = ws.cell(row=row_idx, column=4)
            _, level_fill = get_winrate_style(w["win_rate"])
            level_cell.fill = level_fill
            level_cell.font = Font(name="微软雅黑", size=10, bold=True)

        # 置信度列 (E列)
        conf_cell = ws.cell(row=row_idx, column=5)
        conf = w.get("winrate_confidence", "")
        if conf == "精确":
            conf_cell.font = Font(name="微软雅黑", size=10, bold=True, color="1B5E20")
        elif conf == "高":
            conf_cell.font = Font(name="微软雅黑", size=10, color="2E7D32")
        elif conf == "中":
            conf_cell.font = Font(name="微软雅黑", size=10, color="F57F17")
        elif conf == "低":
            conf_cell.font = Font(name="微软雅黑", size=10, color="B71C1C")

        # GMGN链接 (G列)
        link_cell = ws.cell(row=row_idx, column=7)
        if link_cell.value:
            link_cell.font = Font(color="1565C0", underline="single", size=9)

        # 排名前3标记 (有胜率的)
        if rank <= 3 and w["win_rate"] > 0:
            medal = {1: "🥇", 2: "🥈", 3: "🥉"}
            ws.cell(row=row_idx, column=1).value = f"{medal[rank]} {rank}"
            ws.cell(row=row_idx, column=1).font = Font(name="微软雅黑", size=11, bold=True)

        ws.row_dimensions[row_idx].height = 22

    ws.freeze_panes = "A4"
    if wallets:
        ws.auto_filter.ref = f"A3:{last_col}{3 + len(wallets)}"


def collect_wallet_data():
    """
    主数据收集函数
    从多个维度收集BSC链上的高胜率钱包
    """
    all_wallets = {}

    print("\n" + "=" * 60)
    print("  BSC链高胜率钱包数据采集")
    print("=" * 60)

    # 策略1: 按胜率排名获取 (多时间段)
    for tf in CONFIG["timeframes"]:
        print(f"\n▶ 采集 [{tf}] 胜率排行榜...")
        traders = fetch_top_traders(timeframe=tf, orderby="winrate", limit=CONFIG["top_wallets"])
        for t in traders:
            addr = t.get("wallet_address", "")
            if addr and addr not in all_wallets:
                all_wallets[addr] = t
        time.sleep(CONFIG["request_delay"])

    # 策略2: 按已实现利润排名获取（高利润也可能有高胜率）
    print(f"\n▶ 采集 [7d] 已实现利润排行榜...")
    profit_traders = fetch_top_traders(timeframe="7d", orderby="realized_profit_7d", limit=50)
    for t in profit_traders:
        addr = t.get("wallet_address", "")
        if addr and addr not in all_wallets:
            all_wallets[addr] = t
    time.sleep(CONFIG["request_delay"])

    # 策略3: 按PNL排名
    print(f"\n▶ 采集 [7d] PNL排行榜...")
    pnl_traders = fetch_top_traders(timeframe="7d", orderby="pnl_7d", limit=50)
    for t in pnl_traders:
        addr = t.get("wallet_address", "")
        if addr and addr not in all_wallets:
            all_wallets[addr] = t
    time.sleep(CONFIG["request_delay"])

    print(f"\n✔ 共采集到 {len(all_wallets)} 个不重复钱包地址")

    # 解析所有钱包数据（排行榜API已包含详细字段，无需逐个查询）
    wallets_detailed = []

    print(f"\n▶ 解析钱包数据...")
    for addr, basic_data in all_wallets.items():
        wallet_info = parse_wallet_data(addr, basic_data, None)
        if wallet_info and wallet_info["total_trades"] >= CONFIG["min_trades"]:
            wallets_detailed.append(wallet_info)

    print(f"  已解析 {len(wallets_detailed)} 个有效钱包")

    # 按胜率筛选和排序
    qualified = [
        w for w in wallets_detailed
        if w["win_rate"] >= CONFIG["min_win_rate"]
    ]
    qualified.sort(key=lambda x: (-x["win_rate"], -x["total_profit_usd"]))

    print(f"\n✔ 胜率 ≥ {CONFIG['min_win_rate']}% 的钱包: {len(qualified)} 个")
    return qualified


def parse_wallet_data(address, basic_data, detail_data):
    """解析和标准化钱包数据 (基于GMGN wallets API返回的字段)"""
    try:
        # 交易次数
        buy_count = safe_int(basic_data.get("buy_7d") or basic_data.get("buy", 0))
        sell_count = safe_int(basic_data.get("sell_7d") or basic_data.get("sell", 0))
        total_trades = buy_count + sell_count
        if total_trades == 0:
            total_trades = safe_int(basic_data.get("txs_7d") or basic_data.get("txs", 0))

        # 胜率 (API返回0-1之间的小数)
        win_rate_raw = safe_float(
            basic_data.get("winrate_7d") or basic_data.get("winrate_30d") or basic_data.get("winrate_1d", 0)
        )
        win_rate = win_rate_raw * 100 if win_rate_raw <= 1 else win_rate_raw

        # 已实现利润
        realized_profit = safe_float(
            basic_data.get("realized_profit_7d") or basic_data.get("realized_profit_30d") or basic_data.get("realized_profit_1d", 0)
        )
        unrealized_profit = 0.0  # 排行榜未直接提供，留0

        total_profit = realized_profit
        pnl_7d = safe_float(basic_data.get("pnl_7d", 0))
        pnl_30d = safe_float(basic_data.get("pnl_30d", 0))

        # 平均成本 -> 推算ROI
        avg_cost = safe_float(basic_data.get("avg_cost_7d") or basic_data.get("avg_cost_30d", 0))
        total_cost = avg_cost * buy_count if avg_cost and buy_count else 0
        roi = (realized_profit / total_cost * 100) if total_cost > 0 else (pnl_7d * 100 if pnl_7d else 0)

        # 平均持仓时间 (秒)
        avg_hold_time = basic_data.get("avg_holding_period_7d") or basic_data.get("avg_holding_period_30d", "N/A")

        # 最后活跃时间
        last_active = basic_data.get("last_active", 0)
        if last_active:
            last_active_str = datetime.fromtimestamp(int(last_active)).strftime("%Y-%m-%d %H:%M")
        else:
            last_active_str = "N/A"

        # PNL 分布 (盈利/亏损次数)
        win_count = (
            safe_int(basic_data.get("pnl_lt_2x_num_7d", 0))
            + safe_int(basic_data.get("pnl_2x_5x_num_7d", 0))
            + safe_int(basic_data.get("pnl_gt_5x_num_7d", 0))
        )
        loss_count = (
            safe_int(basic_data.get("pnl_lt_minus_dot5_num_7d", 0))
            + safe_int(basic_data.get("pnl_minus_dot5_0x_num_7d", 0))
        )

        # 标签和社交信息
        tags = basic_data.get("tags", [])
        if isinstance(tags, list):
            tags_str = ", ".join(tags)
        else:
            tags_str = str(tags)

        twitter = basic_data.get("twitter_username", "") or ""
        twitter_name = basic_data.get("twitter_name", "") or basic_data.get("name", "") or ""

        # 交易量
        volume_7d = safe_float(basic_data.get("volume_7d", 0))

        # 关注数
        follow_count = safe_int(basic_data.get("follow_count", 0))

        # 余额
        balance = safe_float(basic_data.get("eth_balance") or basic_data.get("balance", 0))

        # 最大单笔盈利 (从PNL分布推断: 有5x以上的代表大赚)
        pnl_gt_5x = safe_int(basic_data.get("pnl_gt_5x_num_7d", 0))
        pnl_2x_5x = safe_int(basic_data.get("pnl_2x_5x_num_7d", 0))
        max_single_profit = 0  # API未直接提供，标记为0

        # 交易代币数 (buy_count 大致等于代币数)
        token_count = buy_count

        return {
            "address": address,
            "short_address": f"{address[:6]}...{address[-4:]}",
            "win_rate": win_rate,
            "total_trades": total_trades,
            "buy_count": buy_count,
            "sell_count": sell_count,
            "win_count": win_count,
            "loss_count": loss_count,
            "realized_profit_usd": realized_profit,
            "unrealized_profit_usd": unrealized_profit,
            "total_profit_usd": total_profit,
            "pnl_7d": pnl_7d,
            "pnl_30d": pnl_30d,
            "roi_percent": roi,
            "avg_hold_time": format_hold_time(avg_hold_time),
            "max_single_profit": max_single_profit,
            "token_count": token_count,
            "last_active": last_active_str,
            "tags": tags_str,
            "twitter": twitter_name if twitter_name else twitter,
            "follow_count": follow_count,
            "volume_7d": volume_7d,
            "balance_bnb": balance,
            "pnl_gt_5x_count": pnl_gt_5x,
            "pnl_2x_5x_count": pnl_2x_5x,
            "gmgn_link": f"https://gmgn.ai/bsc/address/{address}",
        }
    except Exception as e:
        print(f"\n  [解析错误] {address[:10]}...: {e}")
        return None


# ============================================================
# XLSX 报表生成模块
# ============================================================

def generate_xlsx(wallets, filename):
    """生成带格式的 XLSX 工作表"""
    if not wallets:
        print("\n⚠ 没有符合条件的钱包数据，尝试生成示例报表...")
        wallets = generate_demo_data()

    wb = Workbook()

    # --- Sheet 1: 高胜率钱包总览 ---
    create_overview_sheet(wb, wallets)

    # --- Sheet 2: 胜率分布统计 ---
    create_stats_sheet(wb, wallets)

    # --- Sheet 3: 使用说明 ---
    create_help_sheet(wb)

    # 保存
    filepath = os.path.join(os.path.dirname(os.path.abspath(__file__)), filename)
    wb.save(filepath)
    print(f"\n✔ 报表已保存: {filepath}")
    return filepath


def create_overview_sheet(wb, wallets):
    """创建主数据总览表"""
    ws = wb.active
    ws.title = "高胜率钱包排行"

    # 样式定义
    title_font = Font(name="微软雅黑", size=16, bold=True, color="FFFFFF")
    title_fill = PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid")
    header_font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    data_font = Font(name="Consolas", size=10)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="E0E0E0"),
        right=Side(style="thin", color="E0E0E0"),
        top=Side(style="thin", color="E0E0E0"),
        bottom=Side(style="thin", color="E0E0E0"),
    )

    # 胜率等级颜色
    def get_winrate_style(rate):
        if rate >= 80:
            return Font(name="Consolas", size=11, bold=True, color="1B5E20"), \
                   PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
        elif rate >= 65:
            return Font(name="Consolas", size=11, bold=True, color="33691E"), \
                   PatternFill(start_color="DCEDC8", end_color="DCEDC8", fill_type="solid")
        elif rate >= 50:
            return Font(name="Consolas", size=10, color="F57F17"), \
                   PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
        else:
            return Font(name="Consolas", size=10, color="B71C1C"), \
                   PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")

    def get_profit_font(value):
        if value > 0:
            return Font(name="Consolas", size=10, color="1B5E20")
        elif value < 0:
            return Font(name="Consolas", size=10, color="B71C1C")
        return data_font

    # 标题行
    ws.merge_cells("A1:U1")
    title_cell = ws["A1"]
    title_cell.value = f"BSC链 Meme币 高胜率钱包分析 | 生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    # 摘要行
    ws.merge_cells("A2:U2")
    summary = ws["A2"]
    avg_wr = sum(w["win_rate"] for w in wallets) / len(wallets) if wallets else 0
    top_profit = max((w.get("total_profit_usd", w.get("total_profit_token", 0)) for w in wallets), default=0)
    summary.value = (
        f"共 {len(wallets)} 个钱包 | "
        f"平均胜率: {avg_wr:.1f}% | "
        f"最高总利润: ${top_profit:,.2f} | "
        f"数据源: GMGN.ai"
    )
    summary.font = Font(name="微软雅黑", size=10, color="616161")
    summary.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
    summary.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 25

    # 表头
    headers = [
        ("排名", 6),
        ("钱包地址", 16),
        ("完整地址", 44),
        ("胜率 %", 12),
        ("胜率等级", 12),
        ("总交易次数", 12),
        ("买入次数", 10),
        ("卖出次数", 10),
        ("盈利次数", 10),
        ("亏损次数", 10),
        ("已实现利润 $", 16),
        ("7D PNL", 12),
        ("30D PNL", 12),
        ("ROI %", 12),
        ("平均持仓时间", 14),
        ("7D交易量 $", 16),
        ("BNB余额", 12),
        ("标签", 20),
        ("Twitter", 18),
        ("关注数", 10),
        ("最后活跃", 18),
    ]

    for col_idx, (header_text, width) in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=header_text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[3].height = 30

    # 数据行
    for row_idx, w in enumerate(wallets, 4):
        rank = row_idx - 3

        # 行背景交替色
        row_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        if rank % 2 == 0:
            row_fill = PatternFill(start_color="FAFAFA", end_color="FAFAFA", fill_type="solid")

        row_data = [
            rank,
            w["short_address"],
            w["address"],
            w["win_rate"],
            get_winrate_level(w["win_rate"]),
            w["total_trades"],
            w["buy_count"],
            w["sell_count"],
            w["win_count"],
            w["loss_count"],
            w["realized_profit_usd"],
            w["pnl_7d"],
            w["pnl_30d"],
            w["roi_percent"],
            w["avg_hold_time"],
            w["volume_7d"],
            w["balance_bnb"],
            w["tags"],
            w.get("twitter", ""),
            w.get("follow_count", 0),
            w["last_active"],
        ]

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = center_align
            cell.border = thin_border
            cell.fill = row_fill
            cell.font = data_font

        # 特殊格式: 胜率列
        wr_cell = ws.cell(row=row_idx, column=4)
        wr_font, wr_fill = get_winrate_style(w["win_rate"])
        wr_cell.font = wr_font
        wr_cell.fill = wr_fill
        wr_cell.number_format = '0.0"%"'

        # 胜率等级列颜色
        level_cell = ws.cell(row=row_idx, column=5)
        _, level_fill = get_winrate_style(w["win_rate"])
        level_cell.fill = level_fill
        level_cell.font = Font(name="微软雅黑", size=10, bold=True)

        # 利润列格式
        for col in [11, 16]:  # 已实现利润、7D交易量
            profit_cell = ws.cell(row=row_idx, column=col)
            profit_cell.font = get_profit_font(profit_cell.value or 0)
            profit_cell.number_format = '$#,##0.00'

        # PNL列格式 (小数形式显示为百分比)
        for col in [12, 13]:  # 7D PNL, 30D PNL
            pnl_cell = ws.cell(row=row_idx, column=col)
            pnl_cell.font = get_profit_font(pnl_cell.value or 0)
            pnl_cell.number_format = '0.00%'

        # ROI 格式
        roi_cell = ws.cell(row=row_idx, column=14)
        roi_cell.font = get_profit_font(w["roi_percent"])
        roi_cell.number_format = '0.0"%"'

        # BNB余额格式
        ws.cell(row=row_idx, column=17).number_format = '0.000'

        # 排名前3特殊标记
        if rank <= 3:
            medal = {1: "🥇", 2: "🥈", 3: "🥉"}
            ws.cell(row=row_idx, column=1).value = f"{medal[rank]} {rank}"
            ws.cell(row=row_idx, column=1).font = Font(name="微软雅黑", size=11, bold=True)

        ws.row_dimensions[row_idx].height = 22

    # 冻结窗格
    ws.freeze_panes = "A4"

    # 自动筛选
    if wallets:
        ws.auto_filter.ref = f"A3:U{3 + len(wallets)}"


def create_stats_sheet(wb, wallets):
    """创建统计分析表"""
    ws = wb.create_sheet("胜率分布统计")

    title_font = Font(name="微软雅黑", size=14, bold=True, color="1B5E20")
    header_font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="E0E0E0"),
        right=Side(style="thin", color="E0E0E0"),
        top=Side(style="thin", color="E0E0E0"),
        bottom=Side(style="thin", color="E0E0E0"),
    )

    # 标题
    ws.merge_cells("A1:F1")
    ws["A1"].value = "胜率区间分布统计"
    ws["A1"].font = title_font
    ws["A1"].alignment = center_align
    ws.row_dimensions[1].height = 35

    # 胜率区间统计
    ranges = [
        ("90%-100%", 90, 100),
        ("80%-90%", 80, 90),
        ("70%-80%", 70, 80),
        ("60%-70%", 60, 70),
        ("50%-60%", 50, 60),
    ]

    stats_headers = ["胜率区间", "钱包数量", "占比", "平均利润 $", "平均交易次数", "区间评级"]
    for col, h in enumerate(stats_headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col)].width = 16

    total_count = len(wallets) if wallets else 1
    for row_idx, (label, low, high) in enumerate(ranges, 4):
        in_range = [w for w in wallets if low <= w["win_rate"] < high]
        if label.startswith("90"):
            in_range = [w for w in wallets if w["win_rate"] >= 90]

        count = len(in_range)
        pct = count / total_count * 100 if total_count else 0
        avg_profit = sum(w.get("total_profit_token", w.get("total_profit_usd", 0)) for w in in_range) / count if count else 0
        avg_trades = sum(w.get("total_trade_on_token", w.get("total_trades", 0)) for w in in_range) / count if count else 0

        rating = "⭐" * max(1, min(5, (low - 40) // 10))

        row_data = [label, count, f"{pct:.1f}%", avg_profit, avg_trades, rating]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.alignment = center_align
            cell.border = thin_border
            if col == 4:
                cell.number_format = '$#,##0.00'

    # Top 10 钱包
    ws.merge_cells("A11:F11")
    ws["A11"].value = "Top 10 高胜率钱包"
    ws["A11"].font = title_font
    ws["A11"].alignment = center_align

    top10_headers = ["排名", "钱包地址", "胜率", "总利润", "交易次数", "GMGN链接"]
    for col, h in enumerate(top10_headers, 1):
        cell = ws.cell(row=12, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    for i, w in enumerate(wallets[:10], 1):
        ws.cell(row=12 + i, column=1, value=i).alignment = center_align
        ws.cell(row=12 + i, column=2, value=w["short_address"]).alignment = center_align
        ws.cell(row=12 + i, column=3, value=f"{w['win_rate']:.1f}%").alignment = center_align

        profit_cell = ws.cell(row=12 + i, column=4, value=w.get("total_profit_token", w.get("total_profit_usd", 0)))
        profit_cell.number_format = '$#,##0.00'
        profit_cell.alignment = center_align

        ws.cell(row=12 + i, column=5, value=w.get("total_trade_on_token", w.get("total_trades", 0))).alignment = center_align
        link_cell = ws.cell(row=12 + i, column=6, value=w["gmgn_link"])
        link_cell.font = Font(color="1565C0", underline="single", size=9)
        link_cell.alignment = center_align

        for col in range(1, 7):
            ws.cell(row=12 + i, column=col).border = thin_border


def create_help_sheet(wb):
    """创建使用说明表"""
    ws = wb.create_sheet("使用说明")
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 60

    title_font = Font(name="微软雅黑", size=14, bold=True, color="1B5E20")
    label_font = Font(name="微软雅黑", size=10, bold=True)
    content_font = Font(name="微软雅黑", size=10)

    ws["A1"].value = "BSC高胜率钱包分析报表 - 使用说明"
    ws["A1"].font = title_font
    ws.merge_cells("A1:B1")

    instructions = [
        ("数据源", "GMGN.ai (gmgn.ai) - 智能钱包追踪平台"),
        ("目标链", "BSC (Binance Smart Chain)"),
        ("代币类型", "Meme币 / 内盘代币"),
        ("胜率计算", "精确胜率来自GMGN排行榜数据; 估算胜率基于all_pnl/realized_profit/pnl_30d等指标综合计算"),
        ("胜率来源", "排行榜 = GMGN精确数据 | 估算 = 根据交易盈亏指标推算"),
        ("置信度", "精确 = 排行榜原始数据 | 高 = 交易量大且有资产 | 中 = 一般交易者 | 低 = 交易数据少"),
        ("胜率等级", "S级 ≥80% | A级 ≥65% | B级 ≥50% | C级 <50%"),
        ("颜色标识", "绿色=高胜率(≥80%) | 浅绿=良好(≥65%) | 黄色=中等(≥50%) | 红色=低(<50%)"),
        ("利润颜色", "绿色=盈利 | 红色=亏损"),
        ("筛选功能", "表头支持自动筛选，可按胜率、利润等排序"),
        ("注意事项", "1. 历史胜率不代表未来表现\n2. 需关注钱包是否为合约地址或机器人\n3. 高胜率可能伴随高风险操作"),
        ("免责声明", "本数据仅供参考，不构成投资建议。加密货币交易具有高风险，请谨慎操作。"),
    ]

    for i, (label, content) in enumerate(instructions, 3):
        ws.cell(row=i, column=1, value=label).font = label_font
        ws.cell(row=i, column=2, value=content).font = content_font
        ws.cell(row=i, column=2).alignment = Alignment(wrap_text=True)
        ws.row_dimensions[i].height = 25 if "\n" not in content else 45


# ============================================================
# 工具函数
# ============================================================

def safe_float(val):
    try:
        return float(val) if val else 0.0
    except (ValueError, TypeError):
        return 0.0

def safe_int(val):
    try:
        return int(val) if val else 0
    except (ValueError, TypeError):
        return 0

def format_hold_time(seconds):
    """格式化持仓时间"""
    if not seconds or seconds == "N/A":
        return "N/A"
    try:
        seconds = int(seconds)
        if seconds < 60:
            return f"{seconds}秒"
        elif seconds < 3600:
            return f"{seconds // 60}分钟"
        elif seconds < 86400:
            return f"{seconds // 3600}小时{(seconds % 3600) // 60}分"
        else:
            return f"{seconds // 86400}天{(seconds % 86400) // 3600}小时"
    except (ValueError, TypeError):
        return str(seconds)

def get_winrate_level(rate):
    if rate >= 80:
        return "S级 - 顶级"
    elif rate >= 65:
        return "A级 - 优秀"
    elif rate >= 50:
        return "B级 - 良好"
    else:
        return "C级 - 一般"


def generate_demo_data():
    """生成示例数据（当API无法访问时使用）"""
    import random
    print("\n⚠ 使用示例数据生成演示报表...")

    demo_wallets = []
    for i in range(30):
        addr = "0x" + "".join(random.choices("0123456789abcdef", k=40))
        win_rate = random.uniform(50, 98)
        total_trades = random.randint(10, 500)
        buy_count = random.randint(total_trades // 3, total_trades * 2 // 3)
        sell_count = total_trades - buy_count
        win_count = int(total_trades * win_rate / 100)
        loss_count = total_trades - win_count
        realized = random.uniform(-5000, 50000)
        unrealized = random.uniform(-2000, 15000)

        demo_wallets.append({
            "address": addr,
            "short_address": f"{addr[:6]}...{addr[-4:]}",
            "win_rate": win_rate,
            "total_trades": total_trades,
            "buy_count": buy_count,
            "sell_count": sell_count,
            "win_count": win_count,
            "loss_count": loss_count,
            "realized_profit_usd": realized,
            "unrealized_profit_usd": unrealized,
            "total_profit_usd": realized + unrealized,
            "pnl_7d": random.uniform(-0.5, 2.0),
            "pnl_30d": random.uniform(-0.5, 3.0),
            "roi_percent": random.uniform(-50, 500),
            "avg_hold_time": format_hold_time(random.randint(60, 86400 * 3)),
            "max_single_profit": random.uniform(100, 30000),
            "token_count": random.randint(5, 100),
            "last_active": (datetime.now() - timedelta(hours=random.randint(0, 72))).strftime("%Y-%m-%d %H:%M"),
            "tags": random.choice(["", "kol", "smart_money", "sniper", "top_followed"]),
            "twitter": "",
            "follow_count": random.randint(0, 50000),
            "volume_7d": random.uniform(1000, 500000),
            "balance_bnb": random.uniform(0.01, 50),
            "pnl_gt_5x_count": random.randint(0, 5),
            "pnl_2x_5x_count": random.randint(0, 10),
            "gmgn_link": f"https://gmgn.ai/bsc/address/{addr}",
        })

    demo_wallets.sort(key=lambda x: -x["win_rate"])
    return demo_wallets


# ============================================================
# 主程序入口
# ============================================================

def main():
    print("""
    ╔══════════════════════════════════════════════════╗
    ║   BSC链 Meme币 高胜率钱包分析器 v5.0            ║
    ║   数据源: GMGN.ai | 输出: XLSX                  ║
    ║                                                  ║
    ║   模式1: 全局排行榜 (默认)                       ║
    ║   模式2: 指定CA分析买入者胜率 (内盘数据)         ║
    ╚══════════════════════════════════════════════════╝
    """)

    # 判断运行模式
    token_address = None

    # 支持命令行传入 CA
    if len(sys.argv) > 1:
        arg = sys.argv[1].strip()
        if arg.startswith("0x") and len(arg) == 42:
            token_address = arg.lower()
        elif arg == "--help":
            print("用法:")
            print("  python bsc_wallet_analyzer.py              # 全局排行榜模式")
            print("  python bsc_wallet_analyzer.py <CA地址>     # 指定CA分析模式")
            print("  python bsc_wallet_analyzer.py 0x1234...    # 示例")
            return
    else:
        # 交互模式
        print("请选择运行模式:")
        print("  1. 全局排行榜 - 获取BSC链高胜率钱包排行")
        print("  2. 指定CA分析 - 分析某个代币的买入者胜率")
        print()
        choice = input("输入选择 (1/2, 默认1): ").strip()

        if choice == "2":
            token_address = input("请输入代币合约地址(CA): ").strip().lower()
            if not (token_address.startswith("0x") and len(token_address) == 42):
                print("❌ 无效的合约地址，请输入42位0x开头的地址")
                return

    if token_address:
        # ========== CA分析模式 ==========
        print(f"\n  📍 CA分析模式: {token_address}")
        wallets, token_info = collect_wallet_data_by_ca(token_address)

        symbol = token_info.get("symbol", "UNKNOWN") if token_info else "UNKNOWN"
        output_file = f"BSC_{symbol}_买入者胜率分析.xlsx"

        print("\n" + "=" * 60)
        print("  生成 XLSX 报表")
        print("=" * 60)

        filepath = generate_ca_xlsx(wallets, token_info, token_address, output_file)
    else:
        # ========== 全局排行榜模式 ==========
        print("  📍 全局排行榜模式")
        wallets = collect_wallet_data()
        output_file = CONFIG["output_file"]

        print("\n" + "=" * 60)
        print("  生成 XLSX 报表")
        print("=" * 60)

        filepath = generate_xlsx(wallets, output_file)

    wallet_count = str(len(wallets)) if wallets else "使用示例数据"
    print(f"""
    ╔══════════════════════════════════════════════════╗
    ║   ✅ 分析完成!                                    ║
    ║   文件: {output_file:<41s} ║
    ║   钱包数: {wallet_count:<39s} ║
    ╚══════════════════════════════════════════════════╝
    """)

    return filepath


if __name__ == "__main__":
    main()
